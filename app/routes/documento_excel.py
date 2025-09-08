from flask import render_template, Response, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from db import get_connection

def vista_reportes():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT g.id_grado_grado, g.grado_grupo,
               CONCAT(d.nombres_directriz, ' ', d.apellidos_directriz) AS director
        FROM GRADO_GRUPO g
        JOIN DIRECTRICES d ON g.director_id = d.id_directrices
    """)
    grados = cursor.fetchall()
    conn.close()

    return render_template("reportes_asistencias.html", grados=grados)


def descargar_asistencias():
    grado_id = request.form.get("grado_id")

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # Info del grado y director
    cursor.execute("""
        SELECT g.grado_grupo, 
               CONCAT(d.nombres_directriz, '_', d.apellidos_directriz) AS director
        FROM GRADO_GRUPO g
        JOIN DIRECTRICES d ON g.director_id = d.id_directrices
        WHERE g.id_grado_grado = %s
    """, (grado_id,))
    grado_info = cursor.fetchone()

    # Asistencias
    cursor.execute("""
        SELECT a.id_asistencia, a.fechaHora, a.nombres_asistencia, a.apellidos_asistencia,
               p.documento_persona, p.tipo_personas
        FROM ASISTENCIA a
        JOIN PERSONAS p ON a.persona_id = p.documento_persona
        WHERE p.grado_grupo_id = %s
        ORDER BY a.fechaHora ASC
    """, (grado_id,))
    rows = cursor.fetchall()
    conn.close()

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4CAF50")  # Verde
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Encabezados
    headers = ["ID", "Fecha y Hora", "Nombres", "Apellidos", "Documento", "Tipo"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col)].width = 20  # Ajuste de ancho

    # Filas de datos
    for row in rows:
        ws.append([
            row["id_asistencia"], row["fechaHora"], row["nombres_asistencia"],
            row["apellidos_asistencia"], row["documento_persona"], row["tipo_personas"]
        ])

    # Aplicar bordes y alineación a los datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # Guardar Excel en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{grado_info['director']}_{grado_info['grado_grupo']}.xlsx"

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )
